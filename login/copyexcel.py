import click
from openpyxl import load_workbook
from openpyxl import Workbook
@click.command()
@click.option('--capitalize',default=False)
@click.option('--preservestyles',default=False)
@click.argument('src',type=str,nargs=-1)
@click.argument('dst',type=str,nargs=1)
def cli(capitalize,preservestyles,src,dst):
    if(capitalize & preservestyles):
        nw = Workbook()
        for s in src:
            wb = load_workbook(filename=s)
            nams = wb.get_sheet_names()
            print(nams)
            for sh in nams:
                new=nw.create_sheet()
                ws = wb[sh]
                new.title = sh
                for row in ws.rows:
                    for cell in row:
                        x = '' + cell.column + str(cell.row)
                        new[x] = cell.value.upper()
                        new[x].style = cell.style
    elif(capitalize & (preservestyles==False)):
        nw = Workbook()
        for s in src:
            wb = load_workbook(filename=s)
            nams = wb.get_sheet_names()
            print(nams)
            for sh in nams:
                new=nw.create_sheet()
                ws = wb[sh]
                new.title = sh
                for row in ws.rows:
                    for cell in row:
                        x = '' + cell.column + str(cell.row)
                        new[x] = cell.value.upper()
    elif((capitalize==False) & preservestyles):
        nw = Workbook()
        for s in src:
            wb = load_workbook(filename=s)
            nams = wb.get_sheet_names()
            print(nams)
            for sh in nams:
                new=nw.create_sheet()
                ws = wb[sh]
                new.title = sh
                for row in ws.rows:
                    for cell in row:
                        x = '' + cell.column + str(cell.row)
                        new[x] = cell.value
                        new[x].style = cell.style
    else:
        nw = Workbook()
        for s in src:
            wb = load_workbook(filename=s)
            nams = wb.get_sheet_names()
            print(nams)
            for sh in nams:
                new=nw.create_sheet()
                ws = wb[sh]
                new.title = sh
                for row in ws.rows:
                    for cell in row:
                        x = '' + cell.column + str(cell.row)
                        new[x] = cell.value
            z=nw.get_sheet_by_name(nw.get_sheet_names()[-1])
            nw.remove_sheet(z)
    z=nw.get_sheet_by_name(nw.get_sheet_names()[0])
    nw.remove_sheet(z)
    nw.save(dst)
if __name__=='__main__':
    cli()
